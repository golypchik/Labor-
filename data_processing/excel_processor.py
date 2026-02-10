#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Обработчик Excel файлов
"""

import openpyxl
from pathlib import Path
from datetime import datetime
import sqlite3


class ExcelProcessor:
    """Класс для обработки Excel файлов"""
    
    def __init__(self, session_manager):
        self.session_manager = session_manager
    
    def process_excel_files(self, file_paths):
        """
        Обработка Excel файлов
        
        Структура Excel:
        - Столбец A: общая информация (строка 5 - имя устройства)
        - Столбец B: временные промежутки (начиная со строки 2)
        - Столбец C: температурные значения (начиная со строки 2)
        - Столбец D: значения влажности (начиная со строки 2)
        """
        logger_data = {}
        
        for file_path in file_paths:
            try:
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                worksheet = workbook.active
                
                # Получаем имя устройства из строки 5, столбец A
                device_name = worksheet.cell(row=5, column=1).value
                if not device_name:
                    device_name = Path(file_path).stem
                
                # Извлекаем данные начиная со строки 2
                times = []
                temperatures = []
                humidities = []
                
                row = 2
                while True:
                    time_val = worksheet.cell(row=row, column=2).value
                    temp_val = worksheet.cell(row=row, column=3).value
                    humidity_val = worksheet.cell(row=row, column=4).value
                    
                    # Прекращаем, если нет данных
                    if not time_val and not temp_val and not humidity_val:
                        break
                    
                    if time_val:
                        times.append(time_val)
                    
                    if temp_val is not None:
                        try:
                            temperatures.append(float(temp_val))
                        except (ValueError, TypeError):
                            pass
                    
                    if humidity_val is not None:
                        try:
                            humidities.append(float(humidity_val))
                        except (ValueError, TypeError):
                            pass
                    
                    row += 1
                
                # Сохраняем данные логгера
                if device_name not in logger_data:
                    logger_data[device_name] = {
                        'times': [],
                        'temperatures': [],
                        'humidities': []
                    }
                
                logger_data[device_name]['times'].extend(times)
                logger_data[device_name]['temperatures'].extend(temperatures)
                logger_data[device_name]['humidities'].extend(humidities)
                
                workbook.close()
                
            except Exception as e:
                print(f"Ошибка обработки файла {file_path}: {e}")
                continue
        
        return logger_data
    
    def save_logger_stats(self, logger_data, period_id):
        """Сохранение статистики логгеров в БД"""
        logger_stats_db_path = self.session_manager.get_logger_stats_db_path()
        conn = sqlite3.connect(logger_stats_db_path)
        cursor = conn.cursor()
        
        for logger_name, data in logger_data.items():
            # Извлекаем номер логгера из имени устройства
            logger_number = self.extract_logger_number(logger_name)
            
            # Температура
            if data['temperatures']:
                min_temp = min(data['temperatures'])
                max_temp = max(data['temperatures'])
                avg_temp = sum(data['temperatures']) / len(data['temperatures'])
                
                cursor.execute("""
                    INSERT INTO logger_stats 
                    (period_id, logger_number, data_type, min_value, max_value, avg_value, logger_type)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (period_id, logger_number, 'temperature', min_temp, max_temp, avg_temp, 'internal'))
            
            # Влажность
            if data['humidities']:
                min_hum = min(data['humidities'])
                max_hum = max(data['humidities'])
                avg_hum = sum(data['humidities']) / len(data['humidities'])
                
                cursor.execute("""
                    INSERT INTO logger_stats 
                    (period_id, logger_number, data_type, min_value, max_value, avg_value, logger_type)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (period_id, logger_number, 'humidity', min_hum, max_hum, avg_hum, 'internal'))
        
        conn.commit()
        conn.close()
    
    @staticmethod
    def extract_logger_number(device_name):
        """Извлечение номера логгера из имени устройства"""
        # Пытаемся найти число в имени устройства
        import re
        numbers = re.findall(r'\d+', str(device_name))
        if numbers:
            return numbers[0]
        return str(device_name)
